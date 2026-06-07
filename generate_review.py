"""
Generate an Excel workbook for conference abstract review distribution.

Sheet 1 – Summary: workload calculator (edit yellow cells, rest auto-update).
Sheet 2 – Distribution: randomised assignment of abstracts to SOC members.

With --config it reads
all conference-specific settings (name, raw abstracts file + column map, SOC
members, min reviews, seed, presentation options, output) from the same
conference.yaml that drives the schedule scripts.  Without --config it keeps the
original CLI behaviour.

    python generate_review.py --config conference.yaml
    python generate_review.py --soc 11 --input raw.xlsx ...   # CLI mode

Re-run to regenerate with a new random seed.
"""

import random
import math
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule
from openpyxl.worksheet.datavalidation import DataValidation
import argparse

# Brand string injected into sheet titles; overridden from the YAML (conf_name)
# or left as a neutral default in CLI mode.
CONF_NAME = "Conference"

# ─────────────────────────────────────────────────────────────────────────────
# INPUT READER
# ─────────────────────────────────────────────────────────────────────────────

def _resolve_col(spec, header):
    """Return 0-based column index from a 1-based int or a header name string."""
    if spec is None:
        return None
    try:
        return int(spec) - 1
    except (ValueError, TypeError):
        pass
    if header:
        for i, h in enumerate(header):
            if h is not None and str(h).strip().lower() == str(spec).strip().lower():
                return i
    raise ValueError(f"Column '{spec}' not found in the header row of the input file.")


def read_abstracts(path, id_col=None, title_col=None, text_col=None,
                   name_col=None, gender_col=None, country_col=None,
                   type_col=None, skip_rows=1):
    """
    Read abstracts from an xlsx file.

    All col args accept a 1-based int or a header name string (case-insensitive).
    skip_rows : number of header rows to skip (default 1).

    Returns:
        abstract_ids : list[str]
        metadata     : dict  {abs_id: {title, text, name, gender, country, type}}
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError(f"No active sheet found in {path}")
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header    = all_rows[skip_rows - 1] if skip_rows > 0 and all_rows else None
    data_rows = all_rows[skip_rows:]

    id_idx      = _resolve_col(id_col,      header)
    title_idx   = _resolve_col(title_col,   header)
    text_idx    = _resolve_col(text_col,    header)
    name_idx    = _resolve_col(name_col,    header)
    gender_idx  = _resolve_col(gender_col,  header)
    country_idx = _resolve_col(country_col, header)
    type_idx    = _resolve_col(type_col,    header)

    def _val(row, idx):
        return str(row[idx]).strip() if idx is not None and idx < len(row) and row[idx] is not None else ""

    abstracts = []
    for i, row in enumerate(data_rows, start=1):
        if all(cell is None for cell in row):
            continue
        abs_id = _val(row, id_idx) if id_idx is not None else f"A{i:03d}"
        if not abs_id:
            abs_id = f"A{i:03d}"
        abstracts.append({
            "id":      abs_id,
            "title":   _val(row, title_idx),
            "text":    _val(row, text_idx),
            "name":    _val(row, name_idx),
            "gender":  _val(row, gender_idx),
            "country": _val(row, country_idx),
            "type":    _val(row, type_idx),
        })

    metadata = {a["id"]: {k: a[k] for k in ("title", "text", "name", "gender", "country", "type")}
                for a in abstracts}
    return [a["id"] for a in abstracts], metadata


# ─────────────────────────────────────────────────────────────────────────────
# ABSTRACTS REFERENCE SHEET
# ─────────────────────────────────────────────────────────────────────────────

def build_abstracts_sheet(wb, abstract_ids, metadata):
    ws = wb.create_sheet("Abstracts")
    ws.sheet_view.showGridLines = False

    has_type = any(metadata.get(a, {}).get("type") for a in abstract_ids)
    has_name = any(metadata.get(a, {}).get("name") for a in abstract_ids)

    # Column layout: ID | [Type] | [Author] | Title | Abstract Text
    col = 1
    COL_ID     = col; col += 1
    COL_TYPE   = col if has_type else None;  col += (1 if has_type else 0)
    COL_NAME   = col if has_name else None;  col += (1 if has_name else 0)
    COL_TITLE  = col; col += 1
    COL_TEXT   = col
    n_cols     = col
    last_col   = get_column_letter(n_cols)

    ws.column_dimensions[get_column_letter(COL_ID)].width    = 12
    if COL_TYPE:
        ws.column_dimensions[get_column_letter(COL_TYPE)].width  = 18
    if COL_NAME:
        ws.column_dimensions[get_column_letter(COL_NAME)].width  = 38
    ws.column_dimensions[get_column_letter(COL_TITLE)].width = 38
    ws.column_dimensions[get_column_letter(COL_TEXT)].width  = 80

    ws.merge_cells(f"A1:{last_col}1")
    t = ws["A1"]
    t.value     = f"{CONF_NAME} – Abstract Reference (read-only)"
    t.font      = Font(bold=True, size=13, color="1565C0", name="Calibri")
    t.alignment = CTR
    t.fill      = PatternFill("solid", fgColor="E3F2FD")
    ws.row_dimensions[1].height = 28

    style_header(ws, 2, COL_ID,    "Abstract ID")
    if COL_TYPE:  style_header(ws, 2, COL_TYPE,  "Type")
    if COL_NAME:  style_header(ws, 2, COL_NAME,  "Author")
    style_header(ws, 2, COL_TITLE, "Title")
    style_header(ws, 2, COL_TEXT,  "Abstract Text")
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = ws["A3"]

    WRAP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for ri, abs_id in enumerate(abstract_ids):
        r    = 3 + ri
        fill = LGREY if ri % 2 == 0 else WHITE
        meta = metadata.get(abs_id, {})

        c = ws.cell(row=r, column=COL_ID, value=abs_id)
        c.fill = fill; c.alignment = CTR; c.border = thin_border(); c.font = SMALL

        if COL_TYPE:
            c = ws.cell(row=r, column=COL_TYPE, value=meta.get("type", ""))
            c.fill = fill; c.alignment = CTR; c.border = thin_border(); c.font = SMALL

        if COL_NAME:
            c = ws.cell(row=r, column=COL_NAME, value=meta.get("name", ""))
            c.fill = fill; c.alignment = WRAP_LEFT; c.border = thin_border(); c.font = SMALL

        c = ws.cell(row=r, column=COL_TITLE, value=meta.get("title", ""))
        c.fill = fill; c.alignment = WRAP_LEFT; c.border = thin_border(); c.font = NORMAL

        c = ws.cell(row=r, column=COL_TEXT, value=meta.get("text", ""))
        c.fill = fill; c.alignment = WRAP_LEFT; c.border = thin_border(); c.font = SMALL

        text_len = len(meta.get("text", ""))
        ws.row_dimensions[r].height = max(30, min(150, 15 + text_len // 6))


# ── defaults (overridden via CLI or by editing the constants below) ───────────
DEFAULT_N_SOC        = 13    # number of SOC members available to review
DEFAULT_N_ABSTRACTS  = 80    # expected number of submitted abstracts
DEFAULT_MIN_REVIEWS  = 3     # minimum independent reviews per abstract
DEFAULT_SEED         = None  # None → different random draw each run

# ── colours ───────────────────────────────────────────────────────────────────
YELLOW  = PatternFill("solid", fgColor="FFF9C4")   # editable input cells
BLUE_H  = PatternFill("solid", fgColor="1565C0")   # header background
LGREY   = PatternFill("solid", fgColor="F5F5F5")   # alternating row
WHITE   = PatternFill("solid", fgColor="FFFFFF")
GREEN   = PatternFill("solid", fgColor="E8F5E9")   # summary highlight
ORANGE  = PatternFill("solid", fgColor="FFF3E0")

BLUE_FG  = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
BOLD     = Font(bold=True, name="Calibri", size=11)
NORMAL   = Font(name="Calibri", size=11)
SMALL    = Font(name="Calibri", size=10)

CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT= Alignment(horizontal="right",  vertical="center")

def thin_border():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    s = Side(style="medium", color="555555")
    return Border(left=s, right=s, top=s, bottom=s)


def style_header(ws, row, col, value, width=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill       = BLUE_H
    cell.font       = BLUE_FG
    cell.alignment  = CTR
    cell.border     = thin_border()
    return cell


def style_input(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill      = YELLOW
    cell.font      = BOLD
    cell.alignment = CTR
    cell.border    = thick_border()
    return cell


def style_calc(ws, row, col, value=None, formula=None, fmt=None, fill=None):
    cell = ws.cell(row=row, column=col, value=value if formula is None else formula)
    cell.fill      = fill or WHITE
    cell.font      = NORMAL
    cell.alignment = CTR
    cell.border    = thin_border()
    if fmt:
        cell.number_format = fmt
    return cell


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 1 – SUMMARY / CALCULATOR
# ─────────────────────────────────────────────────────────────────────────────

def build_summary_sheet(ws, n_soc, n_abstracts, min_reviews):
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    # column widths
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 24

    # ── title ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    title = ws["A1"]
    title.value     = f"{CONF_NAME} – Abstract Review Workload Estimator"
    title.font      = Font(bold=True, size=14, color="1565C0", name="Calibri")
    title.alignment = CTR
    title.fill      = PatternFill("solid", fgColor="E3F2FD")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:D2")
    sub = ws["A2"]
    sub.value     = "Edit the yellow input cells — all other values update automatically."
    sub.font      = Font(italic=True, size=10, color="555555", name="Calibri")
    sub.alignment = CTR
    ws.row_dimensions[2].height = 18

    # ── INPUTS ───────────────────────────────────────────────────────────────
    ws.merge_cells("A3:D3")
    sec = ws["A3"]
    sec.value     = "INPUTS"
    sec.font      = BLUE_FG
    sec.fill      = BLUE_H
    sec.alignment = CTR
    ws.row_dimensions[3].height = 22

    inputs = [
        ("Number of SOC members available to review",    n_soc,        "B4",  "integer"),
        ("Number of submitted abstracts",                 n_abstracts,  "B5",  "integer"),
        ("Minimum independent reviews per abstract (N)", min_reviews,  "B6",  "integer"),
    ]
    labels = [
        "Number of SOC members available to review",
        "Number of submitted abstracts",
        "Minimum independent reviews per abstract (N)",
    ]
    defaults = [n_soc, n_abstracts, min_reviews]
    input_rows = [4, 5, 6]

    for r, (label, val) in enumerate(zip(labels, defaults), start=4):
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = BOLD; lc.alignment = LEFT; lc.border = thin_border()
        style_input(ws, r, 2, val)
        # note column
        note_texts = [
            "SOC members who will participate in reviewing",
            "Total abstracts expected / received",
            "Each abstract read by at least this many reviewers",
        ]
        nc = ws.cell(row=r, column=4, value=note_texts[r - 4])
        nc.font = SMALL; nc.alignment = LEFT; nc.fill = ORANGE
        nc.border = thin_border()
        ws.row_dimensions[r].height = 22

    # ── CALCULATED OUTPUTS ───────────────────────────────────────────────────
    ws.merge_cells("A7:D7")
    sec2 = ws["A7"]
    sec2.value     = "CALCULATED OUTPUTS"
    sec2.font      = BLUE_FG
    sec2.fill      = BLUE_H
    sec2.alignment = CTR
    ws.row_dimensions[7].height = 22

    # named ranges via defined names aren't straightforward; use cell refs
    # B4 = n_soc, B5 = n_abstracts, B6 = min_reviews

    calc_rows = [
        ("Total review slots needed  (= abstracts × N)",
         "=B5*B6", None),
        ("Average abstracts per SOC member  (= total slots / members)",
         "=B5*B6/B4", "0.0"),
        ("Minimum abstracts per member  (floor)",
         "=INT(B5*B6/B4)", None),
        ("Extra abstracts to distribute  (remainder)",
         "=MOD(B5*B6,B4)", None),
        ("Members receiving (floor+1) abstracts",
         "=MOD(B5*B6,B4)", None),
        ("Members receiving floor abstracts",
         "=B4-MOD(B5*B6,B4)", None),
    ]

    for i, (label, formula, fmt) in enumerate(calc_rows):
        r = 8 + i
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = NORMAL; lc.alignment = LEFT; lc.border = thin_border()
        lc.fill = GREEN if i == 0 else WHITE

        vc = ws.cell(row=r, column=2, value=formula)
        vc.font = BOLD; vc.alignment = CTR; vc.border = thin_border()
        vc.fill = GREEN if i == 0 else WHITE
        if fmt:
            vc.number_format = fmt

        ws.row_dimensions[r].height = 22

    # merge label columns C & D for calc rows
    for r in range(8, 8 + len(calc_rows)):
        ws.merge_cells(f"C{r}:D{r}")

    # ── QUICK SCENARIO TABLE ─────────────────────────────────────────────────
    r_start = 8 + len(calc_rows) + 2
    ws.merge_cells(f"A{r_start}:D{r_start}")
    sec3 = ws[f"A{r_start}"]
    sec3.value     = "SCENARIO TABLE  (N = minimum reviews per abstract)"
    sec3.font      = BLUE_FG
    sec3.fill      = BLUE_H
    sec3.alignment = CTR
    ws.row_dimensions[r_start].height = 22

    r_start += 1
    headers = ["N (min reviews)", "Total slots", "Avg per member", "Max per member"]
    for ci, h in enumerate(headers, 1):
        style_header(ws, r_start, ci, h)
    ws.row_dimensions[r_start].height = 22

    for ni, n in enumerate([1, 2, 3, 4, 5], 1):
        r = r_start + ni
        fill = LGREY if ni % 2 == 0 else WHITE
        ws.cell(row=r, column=1, value=n).fill = fill
        ws.cell(row=r, column=1).alignment = CTR
        ws.cell(row=r, column=1).border = thin_border()
        # total slots
        c = ws.cell(row=r, column=2, value=f"=B5*{n}")
        c.fill = fill; c.alignment = CTR; c.border = thin_border()
        # avg
        c = ws.cell(row=r, column=3, value=f"=B5*{n}/B4")
        c.fill = fill; c.alignment = CTR; c.border = thin_border()
        c.number_format = "0.0"
        # max (floor+1)
        c = ws.cell(row=r, column=4, value=f"=INT(B5*{n}/B4)+1")
        c.fill = fill; c.alignment = CTR; c.border = thin_border()
        ws.row_dimensions[r].height = 18

    # highlight row matching current N
    # (static; can't do dynamic highlight without VBA — add a note)
    note_r = r_start + 7
    ws.merge_cells(f"A{note_r}:D{note_r}")
    note = ws[f"A{note_r}"]
    note.value     = "★  Row for N=3 highlighted in the Distribution sheet by default.  Change B6 above to update the summary calculations."
    note.font      = Font(italic=True, size=10, color="777777", name="Calibri")
    note.alignment = LEFT

    # ── PRESENTATION TYPE SUMMARY ─────────────────────────────────────────────
    # In the (now single) scores table: PRES_SUGGEST_COL = FLAG_COL + 4
    #   FLAG_COL = last_member_col + 5 = (n_soc+2) + 5 = n_soc + 7
    #   PRES_SUGGEST_COL = n_soc + 11
    # Data rows in Scores Summary: 5 to n_abstracts+4
    scol  = get_column_letter(n_soc + 11)
    sref  = f"'Scores Summary'!{scol}5:{scol}{n_abstracts+4}"

    ps = note_r + 2
    ws.merge_cells(f"A{ps}:C{ps}")
    sec4 = ws[f"A{ps}"]
    sec4.value     = "PRESENTATION TYPE SUMMARY  (live)"
    sec4.font      = BLUE_FG
    sec4.fill      = BLUE_H
    sec4.alignment = CTR
    ws.row_dimensions[ps].height = 22

    ph = ps + 1
    style_header(ws, ph, 1, "# Long oral")
    style_header(ws, ph, 2, "# Short oral")
    style_header(ws, ph, 3, "# Poster")
    ws.row_dimensions[ph].height = 28

    dr = ph + 1
    for ci, pattern in enumerate(["Long oral", "Short oral", "Poster"], start=1):
        c = ws.cell(row=dr, column=ci,
                    value=f'=COUNTIF({sref},"{pattern}")')
        c.font = Font(bold=True, size=14, name="Calibri")
        c.alignment = CTR; c.fill = GREEN; c.border = thin_border()
    ws.row_dimensions[dr].height = 28


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 2 – RANDOM DISTRIBUTION
# ─────────────────────────────────────────────────────────────────────────────

def build_distribution_sheet(ws, soc_members, abstract_ids, min_reviews, seed, metadata=None):
    ws.title = "Distribution"
    ws.sheet_view.showGridLines = False

    n_soc       = len(soc_members)
    n_abstracts = len(abstract_ids)

    if seed is not None:
        random.seed(seed)

    # ── random assignment ────────────────────────────────────────────────────
    # Guarantee each abstract gets exactly min_reviews DISTINCT reviewers.
    # The previous pool-slicing approach could put two copies of the same
    # abstract in one member's consecutive slice; the duplicate collapsed to
    # one ✓, silently dropping that abstract below the min_reviews threshold.

    total_slots   = n_abstracts * min_reviews
    extras_needed = (n_soc - total_slots % n_soc) % n_soc
    per_member    = (total_slots + extras_needed) // n_soc

    # Track how many more times each abstract needs to be assigned.
    # Randomly chosen abstracts get one extra review to balance the workload.
    review_counts = {a: min_reviews for a in abstract_ids}
    if extras_needed:
        for a in random.sample(abstract_ids, extras_needed):
            review_counts[a] += 1

    remaining = dict(review_counts)
    assignments = {m: [] for m in soc_members}

    members_shuffled = soc_members[:]
    random.shuffle(members_shuffled)

    # For each member, pick the per_member abstracts with the most remaining
    # demand (random tiebreak). Keeping counts balanced this way guarantees
    # the final member always finds exactly per_member eligible abstracts.
    for member in members_shuffled:
        eligible = [a for a in abstract_ids if remaining[a] > 0]
        random.shuffle(eligible)                          # random tiebreak
        eligible.sort(key=lambda a: remaining[a], reverse=True)
        chosen = eligible[:per_member]
        for a in chosen:
            remaining[a] -= 1
        assignments[member] = sorted(chosen)

    max_per_member = max(len(v) for v in assignments.values())

    # ── column layout ────────────────────────────────────────────────────────
    # Col A: abstract ID, Col B: abstract title placeholder,
    # Then one column per SOC member (col 3 onwards)
    COL_ABS_ID    = 1   # A
    COL_ABS_TITLE = 2   # B
    FIRST_SOC_COL = 3   # C onwards

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 30
    for ci, _ in enumerate(soc_members, start=FIRST_SOC_COL):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    # ── title ────────────────────────────────────────────────────────────────
    last_col = FIRST_SOC_COL + n_soc - 1
    ws.merge_cells(f"A1:{get_column_letter(last_col)}1")
    title = ws["A1"]
    title.value     = f"{CONF_NAME} – Abstract Review Assignment (random draw)"
    title.font      = Font(bold=True, size=13, color="1565C0", name="Calibri")
    title.alignment = CTR
    title.fill      = PatternFill("solid", fgColor="E3F2FD")
    ws.row_dimensions[1].height = 28

    seed_text = f"Random seed: {seed}" if seed is not None else "Random seed: not fixed (re-run script for a new draw)"
    ws.merge_cells(f"A2:{get_column_letter(last_col)}2")
    sub = ws["A2"]
    sub.value     = (
        f"Abstracts: {n_abstracts}   |   SOC reviewers: {n_soc}   |   "
        f"Min reviews per abstract: {min_reviews}   |   {seed_text}"
    )
    sub.font      = Font(italic=True, size=10, color="555555", name="Calibri")
    sub.alignment = CTR
    ws.row_dimensions[2].height = 18

    # ── TWO sub-tables ───────────────────────────────────────────────────────
    # Table A (rows 4+): abstracts as rows, SOC members as columns (checkbox-style)
    # Table B (further down): SOC members as rows, list of assigned abstracts

    # ── TABLE A: Abstract × Reviewer matrix ─────────────────────────────────
    MATRIX_START = 4

    ws.merge_cells(f"A{MATRIX_START}:{get_column_letter(last_col)}{MATRIX_START}")
    sec = ws[f"A{MATRIX_START}"]
    sec.value     = "TABLE A — Reviewer Matrix  (✓ = assigned)"
    sec.font      = BLUE_FG
    sec.fill      = BLUE_H
    sec.alignment = CTR
    ws.row_dimensions[MATRIX_START].height = 22

    header_row = MATRIX_START + 1
    style_header(ws, header_row, COL_ABS_ID,    "Abstract ID")
    style_header(ws, header_row, COL_ABS_TITLE, "Title / Topic (fill in)")
    for ci, member in enumerate(soc_members, start=FIRST_SOC_COL):
        style_header(ws, header_row, ci, member)
    ws.row_dimensions[header_row].height = 50

    # freeze header row
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    for ri, abs_id in enumerate(abstract_ids):
        r = header_row + 1 + ri
        fill = LGREY if ri % 2 == 0 else WHITE

        c = ws.cell(row=r, column=COL_ABS_ID, value=abs_id)
        c.fill = fill; c.alignment = CTR; c.border = thin_border(); c.font = SMALL

        title_val = metadata.get(abs_id, {}).get("title", "") if metadata else ""
        c = ws.cell(row=r, column=COL_ABS_TITLE, value=title_val)
        c.fill = fill; c.alignment = LEFT; c.border = thin_border()

        for ci, member in enumerate(soc_members, start=FIRST_SOC_COL):
            assigned = abs_id in assignments[member]
            val  = "✓" if assigned else ""
            afill = PatternFill("solid", fgColor="C8E6C9") if assigned else fill
            c = ws.cell(row=r, column=ci, value=val)
            c.fill = afill; c.alignment = CTR; c.border = thin_border()
            c.font = Font(bold=True, color="2E7D32", name="Calibri") if assigned else SMALL

        ws.row_dimensions[r].height = 16

    # totals row
    total_r = header_row + 1 + n_abstracts
    c = ws.cell(row=total_r, column=COL_ABS_ID, value="TOTAL →")
    c.font = BOLD; c.alignment = RIGHT; c.fill = GREEN; c.border = thin_border()
    c = ws.cell(row=total_r, column=COL_ABS_TITLE, value="(abstracts per reviewer)")
    c.font = SMALL; c.fill = GREEN; c.alignment = LEFT; c.border = thin_border()

    for ci, member in enumerate(soc_members, start=FIRST_SOC_COL):
        count = len(assignments[member])
        c = ws.cell(row=total_r, column=ci, value=count)
        c.font = BOLD; c.alignment = CTR; c.fill = GREEN; c.border = thin_border()
    ws.row_dimensions[total_r].height = 20

    # ── TABLE B: per-member list ─────────────────────────────────────────────
    LIST_START = total_r + 3
    ws.merge_cells(f"A{LIST_START}:{get_column_letter(last_col)}{LIST_START}")
    sec2 = ws[f"A{LIST_START}"]
    sec2.value     = "TABLE B — Per-Reviewer Abstract List"
    sec2.font      = BLUE_FG
    sec2.fill      = BLUE_H
    sec2.alignment = CTR
    ws.row_dimensions[LIST_START].height = 22

    # sub-header
    lh = LIST_START + 1
    style_header(ws, lh, 1, "SOC Member")
    style_header(ws, lh, 2, "# Assigned")
    for ci in range(3, 3 + max_per_member):
        style_header(ws, lh, ci, f"Abstract {ci - 2}")
    ws.row_dimensions[lh].height = 22
    # set widths for abstract columns (narrower)
    for ci in range(3, 3 + max_per_member):
        ws.column_dimensions[get_column_letter(ci)].width = max(
            ws.column_dimensions[get_column_letter(ci)].width, 10
        )

    for mi, member in enumerate(soc_members):
        r = lh + 1 + mi
        fill = LGREY if mi % 2 == 0 else WHITE
        c = ws.cell(row=r, column=1, value=member)
        c.font = BOLD; c.alignment = LEFT; c.fill = fill; c.border = thin_border()
        c = ws.cell(row=r, column=2, value=len(assignments[member]))
        c.font = BOLD; c.alignment = CTR; c.fill = GREEN; c.border = thin_border()
        for ci, abs_id in enumerate(assignments[member], start=3):
            c = ws.cell(row=r, column=ci, value=abs_id)
            c.fill = fill; c.alignment = CTR; c.border = thin_border(); c.font = SMALL
        ws.row_dimensions[r].height = 18

    # ── SHEET 3: per-member individual sheets note ────────────────────────────
    return assignments


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 3+: one tab per SOC member
# ─────────────────────────────────────────────────────────────────────────────

PRES_OPTIONS = [
    "Long oral (30 min)",
    "Short oral (15 min)",
    "Poster",
]
PRES_DV_FORMULA = '"' + ",".join(PRES_OPTIONS) + '"'

REQUIRED_FILL = PatternFill("solid", fgColor="FFF8E1")  # warm yellow


def build_member_sheet(wb, member, abstract_ids_assigned, all_abstract_ids, metadata=None):
    ws = wb.create_sheet(title=member[:31])
    ws.sheet_view.showGridLines = False

    # A: Abstract ID  B: Status  C: Title  D: Abstract Text  E: Score  F: Pres Type  G: Comments
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 11
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 13
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 35

    WRAP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
    TEXT_FILL = PatternFill("solid", fgColor="E8F5E9")   # light green — read-only reference

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value     = f"{CONF_NAME} — Abstract Review Sheet: {member}"
    t.font      = Font(bold=True, size=13, color="1565C0", name="Calibri")
    t.alignment = CTR
    t.fill      = PatternFill("solid", fgColor="E3F2FD")
    ws.row_dimensions[1].height = 28

    n_req   = len(abstract_ids_assigned)
    n_total = len(all_abstract_ids)
    ws.merge_cells("A2:G2")
    s = ws["A2"]
    s.value = (f"{n_req} abstracts REQUIRED (highlighted yellow).  "
               f"All {n_total} abstracts shown — you may optionally score any others.  "
               f"Please enter Score (1–10) and Presentation Type for all abstracts you assess.")
    s.font      = Font(italic=True, size=10, color="555555", name="Calibri")
    s.alignment = CTR
    ws.row_dimensions[2].height = 20

    headers = ["Abstract ID", "Status", "Title", "Abstract Text", "Score (1–10)", "Presentation Type", "Comments"]
    for ci, h in enumerate(headers, 1):
        style_header(ws, 3, ci, h)
    ws.row_dimensions[3].height = 22
    ws.freeze_panes = ws["A4"]

    # Data validation: score 1–10 (now col E)
    score_dv = DataValidation(
        type="whole", operator="between", formula1=1, formula2=10,
        showErrorMessage=True,
        error="Please enter a whole number from 1 to 10.",
        errorTitle="Invalid score",
    )
    ws.add_data_validation(score_dv)

    # Data validation: presentation type dropdown (now col F)
    pres_dv = DataValidation(
        type="list", formula1=PRES_DV_FORMULA,
        showErrorMessage=True,
        error="Please select from the dropdown.",
        errorTitle="Invalid selection",
    )
    ws.add_data_validation(pres_dv)

    required_set = set(abstract_ids_assigned)
    last_data_row = 3 + n_total

    for ri, abs_id in enumerate(all_abstract_ids):
        r      = 4 + ri
        is_req = abs_id in required_set
        fill   = REQUIRED_FILL if is_req else (LGREY if ri % 2 == 0 else WHITE)
        meta   = metadata.get(abs_id, {}) if metadata else {}

        for ci in range(1, 8):
            ws.cell(row=r, column=ci).fill   = fill
            ws.cell(row=r, column=ci).border = thin_border()

        c = ws.cell(row=r, column=1, value=abs_id)
        c.alignment = CTR; c.font = SMALL

        sc = ws.cell(row=r, column=2)
        if is_req:
            sc.value = "REQUIRED"
            sc.font  = Font(bold=True, color="E65100", name="Calibri", size=10)
        else:
            sc.value = "optional"
            sc.font  = Font(color="9E9E9E", name="Calibri", size=10)
        sc.alignment = CTR

        tc = ws.cell(row=r, column=3, value=meta.get("title", ""))
        tc.alignment = LEFT; tc.font = SMALL

        text = meta.get("text", "")
        xtc = ws.cell(row=r, column=4, value=text)
        xtc.fill      = TEXT_FILL
        xtc.alignment = WRAP_LEFT
        xtc.font      = Font(name="Calibri", size=9, color="333333")
        xtc.border    = thin_border()

        ws.cell(row=r, column=5).alignment = CTR    # score
        ws.cell(row=r, column=6).alignment = CTR    # pres type
        ws.cell(row=r, column=7).alignment = LEFT   # comments

        # Row height based on text length so it wraps nicely
        # col D is 60 chars wide; ~9pt font fits ~10 chars/line at that width
        chars_per_line = 85
        lines = max(1, math.ceil(len(text) / chars_per_line)) if text else 1
        ws.row_dimensions[r].height = max(30, min(200, lines * 14 + 4))

    score_dv.add(f"E4:E{last_data_row}")
    pres_dv.add(f"F4:F{last_data_row}")

    note_r = last_data_row + 2
    ws.merge_cells(f"A{note_r}:G{note_r}")
    n = ws[f"A{note_r}"]
    n.value = ("Score guide:  10 = exceptional  ·  8–9 = excellent  ·  6–7 = good  ·  "
               "4–5 = acceptable  ·  2–3 = weak  ·  1 = reject")
    n.font      = Font(italic=True, size=9, color="888888", name="Calibri")
    n.alignment = LEFT


# ─────────────────────────────────────────────────────────────────────────────
# SHEET 3 – SCORES SUMMARY (auto-pulls from per-member tabs via INDEX/MATCH)
# ─────────────────────────────────────────────────────────────────────────────

def build_scores_summary_sheet(ws, soc_members, abstract_ids, assignments, metadata=None):
    ws.title = "Scores Summary"
    ws.sheet_view.showGridLines = False

    n_members        = len(soc_members)
    FIRST_MEMBER_COL = 3  # C
    last_member_col  = FIRST_MEMBER_COL + n_members - 1
    AVG_COL          = last_member_col + 1
    CNT_COL          = last_member_col + 2
    MIN_COL          = last_member_col + 3
    MAX_COL          = last_member_col + 4
    FLAG_COL         = last_member_col + 5
    # Presentation type summary appended as extra columns on the same table
    PRES_LONG_COL    = FLAG_COL + 1
    PRES_SHORT_COL   = FLAG_COL + 2
    PRES_POSTER_COL  = FLAG_COL + 3
    PRES_SUGGEST_COL = FLAG_COL + 4
    # Submitter info columns (populated when metadata contains name/gender/country)
    NAME_COL    = PRES_SUGGEST_COL + 1
    GENDER_COL  = PRES_SUGGEST_COL + 2
    COUNTRY_COL = PRES_SUGGEST_COL + 3
    last_col    = COUNTRY_COL

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 32
    for ci in range(FIRST_MEMBER_COL, last_member_col + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 11
    ws.column_dimensions[get_column_letter(AVG_COL)].width          = 13
    ws.column_dimensions[get_column_letter(CNT_COL)].width          = 11
    ws.column_dimensions[get_column_letter(MIN_COL)].width          = 9
    ws.column_dimensions[get_column_letter(MAX_COL)].width          = 9
    ws.column_dimensions[get_column_letter(FLAG_COL)].width         = 16
    ws.column_dimensions[get_column_letter(PRES_LONG_COL)].width    = 12
    ws.column_dimensions[get_column_letter(PRES_SHORT_COL)].width   = 12
    ws.column_dimensions[get_column_letter(PRES_POSTER_COL)].width  = 10
    ws.column_dimensions[get_column_letter(PRES_SUGGEST_COL)].width = 18
    ws.column_dimensions[get_column_letter(NAME_COL)].width         = 22
    ws.column_dimensions[get_column_letter(GENDER_COL)].width       = 10
    ws.column_dimensions[get_column_letter(COUNTRY_COL)].width      = 16

    LC = get_column_letter(last_col)

    # ── title / legend ───────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{LC}1")
    t = ws["A1"]
    t.value     = f"{CONF_NAME} – Abstract Scores & Presentation Type Summary"
    t.font      = Font(bold=True, size=14, color="1565C0", name="Calibri")
    t.alignment = CTR
    t.fill      = PatternFill("solid", fgColor="E3F2FD")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{LC}2")
    s = ws["A2"]
    s.value = ("Scores (1–10) and presentation type recommendations pulled live from reviewer tabs.  "
               "Grey = not required (voluntary score still counts).  "
               "⚠ = scores span ≥ 4 points.  Suggested format = majority vote.")
    s.font      = Font(italic=True, size=10, color="555555", name="Calibri")
    s.alignment = CTR
    ws.row_dimensions[2].height = 18

    ws.merge_cells(f"A3:{LC}3")
    sec = ws["A3"]
    sec.value     = ("SCORES  1=reject · 2–3=weak · 4–5=acceptable · 6–7=good · "
                     "8–9=excellent · 10=exceptional   |   PRESENTATION TYPE  →")
    sec.font      = BLUE_FG
    sec.fill      = BLUE_H
    sec.alignment = CTR
    ws.row_dimensions[3].height = 22

    # ── column headers ───────────────────────────────────────────────────────
    style_header(ws, 4, 1, "Abstract ID")
    style_header(ws, 4, 2, "Title / Topic")
    for ci, member in enumerate(soc_members, start=FIRST_MEMBER_COL):
        style_header(ws, 4, ci, member)
    style_header(ws, 4, AVG_COL,          "Avg Score")
    style_header(ws, 4, CNT_COL,          "# Reviews\nentered")
    style_header(ws, 4, MIN_COL,          "Min")
    style_header(ws, 4, MAX_COL,          "Max")
    style_header(ws, 4, FLAG_COL,         "Flag")
    style_header(ws, 4, PRES_LONG_COL,    "# Long\noral")
    style_header(ws, 4, PRES_SHORT_COL,   "# Short\noral")
    style_header(ws, 4, PRES_POSTER_COL,  "# Poster")
    style_header(ws, 4, PRES_SUGGEST_COL, "Suggested\nFormat")
    style_header(ws, 4, NAME_COL,         "Submitter")
    style_header(ws, 4, GENDER_COL,       "Gender")
    style_header(ws, 4, COUNTRY_COL,      "Country")
    ws.row_dimensions[4].height = 50

    ws.freeze_panes = ws.cell(row=5, column=1)

    NOT_ASSIGNED = PatternFill("solid", fgColor="E0E0E0")
    INDIGO       = PatternFill("solid", fgColor="E8EAF6")
    fm = get_column_letter(FIRST_MEMBER_COL)
    lm = get_column_letter(last_member_col)

    ll = get_column_letter(PRES_LONG_COL)
    sl = get_column_letter(PRES_SHORT_COL)
    pl = get_column_letter(PRES_POSTER_COL)

    for ri, abs_id in enumerate(abstract_ids):
        r    = 5 + ri
        fill = LGREY if ri % 2 == 0 else WHITE

        c = ws.cell(row=r, column=1, value=abs_id)
        c.fill = fill; c.alignment = CTR; c.border = thin_border(); c.font = SMALL
        title_val = metadata.get(abs_id, {}).get("title", "") if metadata else ""
        c = ws.cell(row=r, column=2, value=title_val)
        c.fill = fill; c.alignment = LEFT; c.border = thin_border()

        for ci, member in enumerate(soc_members, start=FIRST_MEMBER_COL):
            safe     = member[:31].replace("'", "''")
            assigned = abs_id in assignments[member]
            formula  = (f"=IFERROR(INDEX('{safe}'!E:E,"
                        f"MATCH(\"{abs_id}\",'{safe}'!A:A,0)),\"\")")
            c = ws.cell(row=r, column=ci, value=formula)
            c.fill      = fill if assigned else NOT_ASSIGNED
            c.font      = Font(bold=True, name="Calibri", size=11) if assigned else SMALL
            c.alignment = CTR; c.border = thin_border()

        rng = f"{fm}{r}:{lm}{r}"

        ac = ws.cell(row=r, column=AVG_COL, value=f'=IFERROR(AVERAGEIF({rng},">0"),"")')
        ac.fill = GREEN; ac.alignment = CTR; ac.border = thin_border()
        ac.font = BOLD; ac.number_format = "0.00"

        cc = ws.cell(row=r, column=CNT_COL, value=f'=COUNTIF({rng},">0")')
        cc.fill = fill; cc.alignment = CTR; cc.border = thin_border(); cc.font = NORMAL

        # MINIFS/MAXIFS get written as _xludf.MINIFS by openpyxl (unrecognised function),
        # which Numbers rejects. Use SMALL+COUNTIF and MAX+COUNTIF instead.
        n_zeros = f'COUNTIF({rng},"<=0")'
        n_pos   = f'COUNTIF({rng},">0")'
        mic = ws.cell(row=r, column=MIN_COL,
                      value=f'=IF({n_pos}=0,"",SMALL({rng},{n_zeros}+1))')
        mic.fill = fill; mic.alignment = CTR; mic.border = thin_border(); mic.font = SMALL

        mxc = ws.cell(row=r, column=MAX_COL,
                      value=f'=IF({n_pos}=0,"",MAX({rng}))')
        mxc.fill = fill; mxc.alignment = CTR; mxc.border = thin_border(); mxc.font = SMALL

        cl  = get_column_letter(CNT_COL)
        mil = get_column_letter(MIN_COL)
        mxl = get_column_letter(MAX_COL)
        # "⚠ Discuss" fires if any reviewer wrote FLAG in comments OR score spread >= 4
        # ISNUMBER(SEARCH(...)) is more compatible than COUNTIFS wildcards across
        # Excel and Numbers; ISNUMBER silently returns FALSE on any error.
        comment_flag = "+".join(
            f"ISNUMBER(SEARCH(\"FLAG\",IFERROR(INDEX("
            f"'{m[:31].replace(chr(39), chr(39)*2)}'!G:G,"
            f"MATCH(\"{abs_id}\",'{m[:31].replace(chr(39), chr(39)*2)}'!A:A,0)),\"\""  ")))"
            for m in soc_members
        )
        # IF short-circuits so MAX-MIN is never evaluated when count=0,
        # avoiding #VALUE! from ""- "" when no scores entered yet
        spread_check = f'IF({cl}{r}>0,{mxl}{r}-{mil}{r}>=4,FALSE)'
        fc = ws.cell(row=r, column=FLAG_COL,
                     value=f'=IF(OR(({comment_flag})>0,{spread_check}),"⚠ Discuss","")')
        fc.fill = fill; fc.alignment = CTR; fc.border = thin_border()
        fc.font = Font(bold=True, color="B71C1C", name="Calibri", size=11)

        # Presentation type counts: COUNTIFS on each reviewer's sheet (col F)
        # COUNTIFS supports wildcards, so "Long oral*" matches all long oral variants
        def pres_formula(pattern):
            parts = [
                f"COUNTIFS('{m[:31].replace(chr(39), chr(39)*2)}'!A:A,"
                f"\"{abs_id}\",'{m[:31].replace(chr(39), chr(39)*2)}'!F:F,\"{pattern}\")"
                for m in soc_members
            ]
            return "=" + "+".join(parts)

        lc = ws.cell(row=r, column=PRES_LONG_COL,   value=pres_formula("Long oral*"))
        sc = ws.cell(row=r, column=PRES_SHORT_COL,  value=pres_formula("Short oral*"))
        pc = ws.cell(row=r, column=PRES_POSTER_COL, value=pres_formula("Poster"))
        for cell in (lc, sc, pc):
            cell.fill = fill; cell.alignment = CTR
            cell.border = thin_border(); cell.font = NORMAL

        suggest = (f'=IF(AND({ll}{r}=0,{sl}{r}=0,{pl}{r}=0),"",'
                   f'IF(AND({ll}{r}>=MAX({ll}{r},{sl}{r},{pl}{r})),"Long oral",'
                   f'IF({sl}{r}>=MAX({sl}{r},{pl}{r}),"Short oral","Poster")))')
        sg = ws.cell(row=r, column=PRES_SUGGEST_COL, value=suggest)
        sg.fill = INDIGO; sg.alignment = CTR; sg.border = thin_border()
        sg.font = Font(bold=True, color="283593", name="Calibri", size=10)

        # Submitter info — static values from metadata
        meta = metadata.get(abs_id, {}) if metadata else {}
        PALE = PatternFill("solid", fgColor="F3E5F5")
        for col_idx, key in [(NAME_COL, "name"), (GENDER_COL, "gender"), (COUNTRY_COL, "country")]:
            c = ws.cell(row=r, column=col_idx, value=meta.get(key, ""))
            c.fill = PALE; c.alignment = CTR; c.border = thin_border(); c.font = SMALL

        ws.row_dimensions[r].height = 18

    # ── totals row ───────────────────────────────────────────────────────────
    tr = 5 + len(abstract_ids)
    ws.merge_cells(f"A{tr}:B{tr}")
    c = ws.cell(row=tr, column=1, value="TOTALS / AVERAGES →")
    c.font = BOLD; c.alignment = RIGHT; c.fill = GREEN; c.border = thin_border()

    for ci in range(FIRST_MEMBER_COL, last_member_col + 1):
        cl2 = get_column_letter(ci)
        c = ws.cell(row=tr, column=ci,
                    value=f'=IFERROR(AVERAGEIF({cl2}5:{cl2}{tr-1},">0"),"")')
        c.fill = GREEN; c.alignment = CTR; c.border = thin_border()
        c.font = BOLD; c.number_format = "0.00"

    acl = get_column_letter(AVG_COL)
    c = ws.cell(row=tr, column=AVG_COL,
                value=f'=IFERROR(AVERAGEIF({acl}5:{acl}{tr-1},">0"),"")')
    c.fill = PatternFill("solid", fgColor="A5D6A7"); c.alignment = CTR
    c.border = thin_border()
    c.font = Font(bold=True, size=12, name="Calibri"); c.number_format = "0.00"

    ccl = get_column_letter(CNT_COL)
    c = ws.cell(row=tr, column=CNT_COL, value=f"=SUM({ccl}5:{ccl}{tr-1})")
    c.fill = GREEN; c.alignment = CTR; c.border = thin_border(); c.font = BOLD

    flagcl = get_column_letter(FLAG_COL)
    c = ws.cell(row=tr, column=FLAG_COL,
                value=f'=(COUNTIF({flagcl}5:{flagcl}{tr-1},"⚠ Discuss")'
                      f'+COUNTIF({flagcl}5:{flagcl}{tr-1},"🚩 FLAG"))&" flagged"')
    c.fill = PatternFill("solid", fgColor="FFCDD2"); c.alignment = CTR
    c.border = thin_border()
    c.font = Font(bold=True, color="B71C1C", name="Calibri", size=10)

    for col_idx, col_let in [(PRES_LONG_COL, ll), (PRES_SHORT_COL, sl), (PRES_POSTER_COL, pl)]:
        c = ws.cell(row=tr, column=col_idx,
                    value=f"=SUM({col_let}5:{col_let}{tr-1})")
        c.fill = GREEN; c.alignment = CTR; c.border = thin_border(); c.font = BOLD

    sgl = get_column_letter(PRES_SUGGEST_COL)
    c = ws.cell(row=tr, column=PRES_SUGGEST_COL,
                value=(f'=COUNTIF({sgl}5:{sgl}{tr-1},"Long oral")&" long  /  "'
                       f'&COUNTIF({sgl}5:{sgl}{tr-1},"Short oral")&" short  /  "'
                       f'&COUNTIF({sgl}5:{sgl}{tr-1},"Poster")&" poster"'))
    c.fill = PatternFill("solid", fgColor="C5CAE9"); c.alignment = CTR
    c.border = thin_border()
    c.font = Font(bold=True, color="283593", name="Calibri", size=10)

    # Submitter info totals: unique-value counts computed from metadata
    if metadata:
        names    = [metadata.get(a, {}).get("name",    "") for a in abstract_ids]
        genders  = [metadata.get(a, {}).get("gender",  "") for a in abstract_ids]
        countries= [metadata.get(a, {}).get("country", "") for a in abstract_ids]
        PALE = PatternFill("solid", fgColor="F3E5F5")
        for col_idx, values, label in [
            (NAME_COL,    names,     "submitters"),
            (GENDER_COL,  genders,   "genders"),
            (COUNTRY_COL, countries, "countries"),
        ]:
            filled   = [v for v in values if v]
            unique   = len(set(filled))
            summary  = f"{unique} unique {label}" if filled else ""
            c = ws.cell(row=tr, column=col_idx, value=summary)
            c.fill = PALE; c.alignment = CTR; c.border = thin_border()
            c.font = Font(italic=True, size=10, name="Calibri")

    ws.row_dimensions[tr].height = 22


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def _parse_col(val):
    """Accept '3' → 3 (int) or 'Title' → 'Title' (str) for column specs."""
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return val


def _slug(text):
    """Filesystem-friendly short slug from the conference name."""
    import re as _re
    s = _re.sub(r"[^A-Za-z0-9]+", "_", str(text)).strip("_")
    return s or "conference"


def main():
    parser = argparse.ArgumentParser(description="Generate a conference abstract review Excel workbook.")
    parser.add_argument("--config", default=None,
                        help="Path to conference YAML. When given, the review: section "
                             "supplies all settings below (CLI flags are ignored).")
    parser.add_argument("--soc",       type=int,  default=DEFAULT_N_SOC,       help="Number of SOC members")
    parser.add_argument("--abstracts", type=int,  default=DEFAULT_N_ABSTRACTS, help="Number of abstracts (ignored when --input is used)")
    parser.add_argument("--min",       type=int,  default=DEFAULT_MIN_REVIEWS, help="Min reviews per abstract")
    parser.add_argument("--seed",      type=int,  default=DEFAULT_SEED,        help="Random seed (omit for random)")
    parser.add_argument("--members",   nargs="+", default=None,                help="SOC member names (space separated)")
    parser.add_argument("--out",       default=None,                           help="Output filename")
    # Abstract data input
    parser.add_argument("--input",        default=None, help="Path to xlsx containing abstract data")
    parser.add_argument("--id-col",       default=None, help="Column for abstract ID — 1-based index or header name (default: auto-number)")
    parser.add_argument("--title-col",    default="1",  help="Column for title — 1-based index or header name (default: 1)")
    parser.add_argument("--text-col",     default=None, help="Column for abstract text — 1-based index or header name (optional)")
    parser.add_argument("--name-col",     default=None, help="Column for submitter name (optional)")
    parser.add_argument("--gender-col",   default=None, help="Column for submitter gender (optional)")
    parser.add_argument("--country-col",  default=None, help="Column for submitter country (optional)")
    parser.add_argument("--type-col",     default=None, help="Column for contribution type — 1-based index or header name (optional)")
    parser.add_argument("--review-types", nargs="+",   default=None,
                        help="Contribution type(s) to include in review distribution (case-insensitive). "
                             "All abstracts still appear on the Abstracts sheet. "
                             "Example: --review-types Talk \"Invited talk\"")
    parser.add_argument("--skip-rows",    type=int, default=1, help="Header rows to skip in input file (default: 1)")
    args = parser.parse_args()

    global CONF_NAME, PRES_OPTIONS, PRES_DV_FORMULA

    # ── Resolve settings: YAML (--config) takes precedence over CLI flags ──────
    if args.config:
        from schedule_config import load
        conf = load(args.config)
        rv = conf.REVIEW or {}
        CONF_NAME = conf.CONF_NAME
        pres = rv.get("presentation_options")
        if pres:
            PRES_OPTIONS    = list(pres)
            PRES_DV_FORMULA = '"' + ",".join(PRES_OPTIONS) + '"'

        soc_members = list(rv.get("soc_members") or [])
        if not soc_members:
            raise SystemExit(f"{args.config}: review.soc_members must list the SOC members")
        n_soc        = len(soc_members)
        min_reviews  = int(rv.get("min_reviews", DEFAULT_MIN_REVIEWS))
        seed         = rv.get("seed")
        skip_rows    = int(rv.get("skip_rows", 1))
        review_types = rv.get("review_types")
        cols         = rv.get("columns", {}) or {}
        input_path   = rv.get("raw_abstracts")
        if not input_path:
            raise SystemExit(f"{args.config}: review.raw_abstracts (raw submissions xlsx) is required")
        id_spec      = cols.get("id")
        title_spec   = cols.get("title") or 1
        text_spec    = cols.get("text")
        name_spec    = cols.get("name")
        gender_spec  = cols.get("gender")
        country_spec = cols.get("country")
        type_spec    = cols.get("type")
        out_path     = rv.get("output")
    else:
        n_soc        = args.soc
        min_reviews  = args.min
        seed         = args.seed
        skip_rows    = args.skip_rows
        review_types = args.review_types
        if args.members:
            soc_members = args.members
            n_soc = len(soc_members)
        else:
            soc_members = [f"SOC Member {i+1:02d}" for i in range(n_soc)]
        input_path   = args.input
        id_spec      = _parse_col(args.id_col)
        title_spec   = _parse_col(args.title_col) if args.title_col else 1
        text_spec    = _parse_col(args.text_col)
        name_spec    = _parse_col(args.name_col)
        gender_spec  = _parse_col(args.gender_col)
        country_spec = _parse_col(args.country_col)
        type_spec    = _parse_col(args.type_col)
        out_path     = args.out

    # ── Abstract IDs and metadata ─────────────────────────────────────────────
    metadata = None
    if input_path:
        all_abstract_ids, metadata = read_abstracts(
            input_path,
            id_col      = id_spec,
            title_col   = title_spec,
            text_col    = text_spec,
            name_col    = name_spec,
            gender_col  = gender_spec,
            country_col = country_spec,
            type_col    = type_spec,
            skip_rows   = skip_rows,
        )
        print(f"  Input file  : {input_path}  ({len(all_abstract_ids)} abstracts read)")
        # Filter to requested contribution types for review; keep all for Abstracts sheet
        if review_types:
            allowed = {t.lower() for t in review_types}
            abstract_ids = [a for a in all_abstract_ids
                            if metadata.get(a, {}).get("type", "").lower() in allowed]
            print(f"  Type filter : {review_types}  ({len(abstract_ids)} abstracts for review)")
        else:
            abstract_ids = all_abstract_ids
    else:
        all_abstract_ids = [f"A{i+1:03d}" for i in range(args.abstracts)]
        abstract_ids     = all_abstract_ids

    n_abstracts = len(abstract_ids)

    out_path = out_path or (
        f"{_slug(CONF_NAME)}_abstract_review_N{n_abstracts}_reviewers{n_soc}_min{min_reviews}.xlsx"
    )

    wb = Workbook()
    wb.calculation.calcMode      = "auto"
    wb.calculation.fullCalcOnLoad = True

    # Sheet 1 – Summary
    ws_sum = wb.active
    build_summary_sheet(ws_sum, n_soc, n_abstracts, min_reviews)

    # Sheet 2 – Distribution
    ws_dist = wb.create_sheet("Distribution")
    assignments = build_distribution_sheet(
        ws_dist, soc_members, abstract_ids, min_reviews, seed, metadata=metadata
    )

    # Sheet 3 – Abstracts reference (only when input data was provided; uses full list)
    if metadata:
        build_abstracts_sheet(wb, all_abstract_ids, metadata)

    # Sheet 4 – Scores Summary (created before per-member sheets so
    # INDEX/MATCH formulas can reference the tabs that follow)
    ws_scores = wb.create_sheet("Scores Summary")
    build_scores_summary_sheet(
        ws_scores, soc_members, abstract_ids, assignments, metadata=metadata
    )

    # Sheets 5+ – per-member
    for member in soc_members:
        build_member_sheet(wb, member, assignments[member], abstract_ids, metadata=metadata)

    wb.save(out_path)
    print(f"Saved: {out_path}")
    print(f"  SOC members : {n_soc}")
    print(f"  Abstracts   : {n_abstracts}")
    print(f"  Min reviews : {min_reviews}")
    total = n_abstracts * min_reviews
    avg   = total / n_soc
    print(f"  Total slots : {total}")
    print(f"  Avg/member  : {avg:.1f}  (range: {math.floor(avg)}–{math.ceil(avg)})")


if __name__ == "__main__":
    main()
