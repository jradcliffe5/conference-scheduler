#!/usr/bin/env python3
"""
make_draft_emails.py  —  Draft participant outcome emails from a CONFIRMED schedule.

Run this AFTER the schedule grid is finalised.  For every abstract submitter it
works out the outcome (oral talk with day/time/session, poster, or not accepted),
renders a personalised message, and creates an email DRAFT.  It never sends.

Two modes:
  --dry-run (default)  : write one .eml preview per participant to a folder and
                         print a summary table.  Touches nothing online.
  --create-drafts      : append the messages to your Gmail Drafts folder over IMAP
                         (using an app password from the env var in emails.password_env).

Usage:
    python make_draft_emails.py --config conference.yaml                 # preview
    DRAFT_EMAIL_PASSWORD=<app-pw> \\
        python make_draft_emails.py --config conference.yaml --create-drafts

The app password is a Google "App Password" (Account → Security → App passwords),
NOT your normal login password.  Drafts can be reviewed/edited/sent from any client.
"""

import argparse
import imaplib
import os
import re
import sys
import time
from datetime import date, timedelta
from email.message import EmailMessage
from email.utils import formatdate

import openpyxl

from schedule_config import load, Conf
import fill_schedule as fill


# Require a dotted domain so "wbrisken@nrao" is flagged as needs-review.
EMAIL_RE      = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")
EMAIL_LOOSE   = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+")


def extract_email(submitter: str) -> tuple[str | None, bool]:
    """Return (email, ok). ok=False when missing or clearly malformed."""
    m = EMAIL_RE.search(submitter or "")
    if m:
        return m.group(0), True
    m2 = EMAIL_LOOSE.search(submitter or "")
    if m2:
        return m2.group(0), False   # found something @-ish but no valid domain
    return None, False


def greeting_name(submitter: str) -> str:
    """Best-effort personal name from the 'Name, email, affiliation' blob."""
    first = (submitter or "").split("\n")[0]
    name  = re.split(r"[,\n]", first)[0].strip()
    if not name or EMAIL_LOOSE.search(name):
        return "Colleague"
    return name


def _weekday(iso: str) -> str:
    return date.fromisoformat(iso).strftime("%A %d %b %Y")


# ─── Parse the confirmed schedule into per-abstract outcomes ───────────────────

def parse_schedule(conf: Conf) -> dict[str, dict]:
    """
    Read the finalised schedule workbook and return {abstract_id: outcome dict}.
    outcome kind is 'oral' (with day/time/session) or 'poster'.
    """
    wb   = openpyxl.load_workbook(conf.SCHEDULE_FILE, data_only=True)
    ws   = wb["Programme"]
    meta = wb["_META"]

    sessions: list[tuple] = []     # (session_index, label, first_row, last_row)
    day_col: dict[int, str] = {}   # excel column -> iso date
    for r in meta.iter_rows(values_only=True):
        if r[0] is None:
            continue
        c1 = str(r[1]) if r[1] is not None else ""
        if re.match(r"\d{4}-\d{2}-\d{2}", c1):           # day row
            day_col[int(r[2])] = c1
        elif r[2] is not None and r[3] is not None:       # session row
            sessions.append((int(r[0]), c1, int(r[2]), int(r[3])))

    id_re = re.compile(conf.ID_PATTERN.strip("^$"))   # search anywhere in cell text
    type_word = {"INV": "invited talk", "LONG": "contributed (long) oral",
                 "SHORT": "contributed (short) oral"}

    outcomes: dict[str, dict] = {}

    # Oral placements from the Programme grid.
    for row in ws.iter_rows():
        for cell in row:
            if not cell.value:
                continue
            txt = str(cell.value)
            m = id_re.search(txt)
            if not m or cell.column not in day_col:
                continue
            aid = m.group(0)
            sess = next((s for s in sessions if s[2] <= cell.row <= s[3]), None)
            if sess is None:
                continue
            si, label, first_row, _ = sess
            slot = cell.row - first_row
            _, sh, sm, _, _ = conf.SESSION_BLOCKS[si]
            start = sh * 60 + sm + slot * conf.SLOT_MIN
            lbl = re.search(r"\[(INV|LONG|SHORT)", txt)
            outcomes[aid] = {
                "kind":    "oral",
                "type":    type_word.get(lbl.group(1) if lbl else "", "oral presentation"),
                "day":     _weekday(day_col[cell.column]),
                "time":    f"{start // 60:02d}:{start % 60:02d}",
                "session": label,
            }

    # Posters / sparklers from the Posters sheet.
    sp_when = ""
    if conf.SPARKLER_SESSIONS:
        di, si = conf.SPARKLER_SESSIONS[0][0], conf.SPARKLER_SESSIONS[0][1]
        d  = conf.CONF_START + timedelta(days=di)
        sl = conf.SESSION_BLOCKS[si]
        sp_when = f"{d.strftime('%A %d %b')} ({sl[0]}, {sl[1]:02d}:{sl[2]:02d}–{sl[3]:02d}:{sl[4]:02d})"
    if "Posters" in wb.sheetnames:
        wp = wb["Posters"]
        for r in wp.iter_rows(min_row=2, values_only=True):
            if not r or not r[0]:
                continue
            aid = str(r[0]).strip()
            if id_re.fullmatch(aid) and aid not in outcomes:
                outcomes[aid] = {"kind": "poster", "when": sp_when}
    return outcomes


# ─── Compose outcome text + email ──────────────────────────────────────────────

def outcome_text(conf: Conf, o: dict | None) -> str:
    if o is None:
        return ("After careful review by the SOC, we are sorry to say that we were "
                "unable to include your contribution in the final programme this time. "
                "We received many more strong submissions than we had slots for, and "
                "we very much hope you will still join us at the meeting.")
    if o["kind"] == "oral":
        article = "an" if o["type"][:1].lower() in "aeiou" else "a"
        return (f"We are pleased to offer your contribution {article} {o['type']}.\n\n"
                f"It is scheduled for {o['day']} at {o['time']}, in the "
                f"\"{o['session']}\" session. Please let us know as soon as possible "
                f"if you are unable to present at this time.")
    when = f" The poster session takes place on {o['when']}." if o.get("when") else ""
    return ("We are pleased to offer your contribution a poster presentation."
            + when +
            "\n\nPosters will be on display throughout the meeting, with a short "
            "\"sparkler\" slot to advertise them to all participants.")


def render_email(conf: Conf, talk, o: dict | None) -> tuple[str, str]:
    em = conf.EMAILS
    ctx = {
        "conf_name":     conf.CONF_NAME,
        "conf_venue":    conf.CONF_VENUE,
        "abstract_id":   talk.abstract_id,
        "title":         talk.title,
        "greeting_name": greeting_name(talk.submitter),
    }
    subject = em.get("subject",
                     "{conf_name}: outcome of your abstract submission ({abstract_id})").format(**ctx)
    signature = em.get("signature", "With best wishes,\nThe {conf_name} SOC").format(**ctx)
    body = (
        f"Dear {ctx['greeting_name']},\n\n"
        f"Thank you for submitting \"{ctx['title']}\" ({ctx['abstract_id']}) to "
        f"{ctx['conf_name']}.\n\n"
        f"{outcome_text(conf, o)}\n\n"
        f"{signature}\n"
    )
    return subject, body


def build_message(conf: Conf, to_addr: str, subject: str, body: str,
                  needs_review: bool) -> EmailMessage:
    em = conf.EMAILS
    msg = EmailMessage()
    from_name = em.get("from_name", conf.CONF_NAME)
    from_addr = em.get("from_address", "")
    msg["From"] = f"{from_name} <{from_addr}>" if from_addr else from_name
    msg["To"]   = to_addr or ""
    if em.get("reply_to"):
        msg["Reply-To"] = em["reply_to"]
    msg["Subject"] = ("[CHECK ADDRESS] " + subject) if needs_review else subject
    msg["Date"] = formatdate(localtime=True)
    msg.set_content(body)
    return msg


# ─── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Draft participant outcome emails from a confirmed schedule (never sends).")
    parser.add_argument("--config", default="conference.yaml")
    parser.add_argument("--create-drafts", action="store_true",
                        help="Append to Gmail Drafts over IMAP (default: dry-run .eml previews)")
    parser.add_argument("--outdir", default="draft_emails",
                        help="Folder for .eml previews in dry-run mode (default: draft_emails)")
    args = parser.parse_args()

    conf = load(args.config)
    talks = fill.read_abstracts(conf)                 # every submitter (master list)
    outcomes = parse_schedule(conf)                   # confirmed placements

    # De-duplicate submitters by abstract id; keep first occurrence.
    seen: set[str] = set()
    recipients = []
    counts = {"oral": 0, "poster": 0, "not accepted": 0, "needs_review": 0, "no_email": 0}
    for t in talks:
        if t.abstract_id in seen:
            continue
        seen.add(t.abstract_id)
        o = outcomes.get(t.abstract_id)
        kind = o["kind"] if o else "not accepted"
        email, ok = extract_email(t.submitter)
        subject, body = render_email(conf, t, o)
        needs_review = (email is None) or (not ok)
        counts[kind if kind in counts else "not accepted"] += 1
        if email is None:
            counts["no_email"] += 1
        elif not ok:
            counts["needs_review"] += 1
        recipients.append({
            "aid": t.abstract_id, "name": greeting_name(t.submitter),
            "email": email or "", "ok": ok and email is not None,
            "kind": kind, "subject": subject, "body": body, "needs_review": needs_review,
        })

    print(f"Schedule   : {conf.SCHEDULE_FILE}")
    print(f"Abstracts  : {conf.INPUT_ABSTRACTS}")
    print(f"Participants: {len(recipients)}  "
          f"(oral {counts['oral']}, poster {counts['poster']}, "
          f"not accepted {counts['not accepted']})")
    flagged = counts["needs_review"] + counts["no_email"]
    if flagged:
        print(f"  ⚠ {flagged} address(es) need checking "
              f"({counts['no_email']} missing, {counts['needs_review']} malformed) "
              f"— their drafts are prefixed [CHECK ADDRESS]")

    if not args.create_drafts:
        os.makedirs(args.outdir, exist_ok=True)
        for r in recipients:
            msg = build_message(conf, r["email"], r["subject"], r["body"], r["needs_review"])
            safe = re.sub(r"[^A-Za-z0-9]+", "_", r["name"])[:40].strip("_") or "participant"
            path = os.path.join(args.outdir, f"{r['aid']}_{safe}.eml")
            with open(path, "wb") as fh:
                fh.write(bytes(msg))
        print(f"\n[dry-run] Wrote {len(recipients)} .eml previews to {args.outdir}/")
        print("Review them, then re-run with --create-drafts to append to Gmail Drafts.")
        for r in recipients[:12]:
            flag = " [CHECK ADDRESS]" if r["needs_review"] else ""
            print(f"  {r['aid']:<6} {r['kind']:<12} {r['email'] or '(no email)':<32} {r['name']}{flag}")
        if len(recipients) > 12:
            print(f"  ... and {len(recipients) - 12} more")
        return

    # ── Create real drafts over IMAP (never sends) ────────────────────────────
    em = conf.EMAILS
    pw_env = em.get("password_env", "DRAFT_EMAIL_PASSWORD")
    password = os.environ.get(pw_env)
    if not password:
        sys.exit(f"\nIMAP app password not found in ${pw_env}.\n"
                 f"Set it, e.g.:  export {pw_env}='your-16-char-app-password'\n"
                 f"(Create one at Google Account → Security → App passwords.)")
    host   = em.get("imap_host", "imap.gmail.com")
    port   = int(em.get("imap_port", 993))
    folder = em.get("drafts_folder", "[Gmail]/Drafts")
    user   = em.get("from_address", "")

    print(f"\nConnecting to {host}:{port} as {user} …")
    imap = imaplib.IMAP4_SSL(host, port)
    imap.login(user, password)
    made = 0
    for r in recipients:
        msg = build_message(conf, r["email"], r["subject"], r["body"], r["needs_review"])
        imap.append(folder, "(\\Draft)", imaplib.Time2Internaldate(time.time()),
                    bytes(msg))
        made += 1
    imap.logout()
    print(f"Created {made} drafts in '{folder}'.  Nothing was sent.")
    print("Open your Drafts folder to review, edit, and send when ready.")


if __name__ == "__main__":
    main()
