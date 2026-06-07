#!/usr/bin/env python3
"""
fill_schedule.py  —  Fill the schedule grid with scored abstracts.

All conference-specific settings come
from a YAML file (default conference.yaml), and spreadsheet columns are located by
HEADER NAME (not fixed position), so any equivalently-structured "Scores ranked"
workbook works without code edits.

    python fill_schedule.py                      # uses conference.yaml
    python fill_schedule.py --config other.yaml
    python fill_schedule.py --dry-run            # plan only, no writes

Rules (durations/types come from the YAML format_map):
  - Invited/Long  : placed FIRST in a session
  - Short oral    : fills remaining session time
  - Posters       : listed on the Posters sheet (sparkler session)
  - Station-tagged high-score talks are grouped into dedicated session(s)
  - NEVER overwrites a cell that already has content.
"""

import argparse
import random
import re
import sys
import textwrap
from dataclasses import dataclass
from datetime import date, time, datetime, timedelta
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from schedule_config import load, Conf


@dataclass
class Talk:
    abstract_id:      str
    title:            str
    submitter:        str
    avg_score:        float
    suggested_format: str
    talk_type:        str   # resolved: "Invited" / "Long" / "Short" / "Poster"
    duration:         int   # minutes
    topic:            str = ""   # tag from the topic_tag column (e.g. "Station")


@dataclass
class SessionSlot:
    """Represents one talk-row cell inside a session for one day."""
    row: int
    col: int
    day_index: int
    session_index: int
    slot_index: int     # 0-based within the session


# ─── Header-based column detection (the genericity layer) ──────────────────────

def _norm(s) -> str:
    """Normalise a header cell: strip, lower-case, collapse internal whitespace."""
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip().lower())


def find_columns(rows, aliases, required, sheet_name, max_scan=15):
    """
    Locate the header row and map logical fields -> 0-based column indices.

    rows     : materialised list of row tuples (values_only).
    aliases  : {field: [acceptable header strings]} (case-insensitive).
    required : fields that MUST be found, else a clear error is raised.
    Returns  : (header_row_index, {field: col_index}).
    """
    id_aliases = {_norm(a) for a in aliases.get("id", [])}
    header_idx = None
    for i, row in enumerate(rows[:max_scan]):
        present = {_norm(v) for v in row if v is not None}
        if id_aliases & present:
            header_idx = i
            break
    if header_idx is None:
        raise ValueError(
            f"Sheet '{sheet_name}': could not find a header row containing any of "
            f"{aliases.get('id')} in the first {max_scan} rows.")

    header = rows[header_idx]
    norm_to_idx: dict[str, int] = {}
    for idx, v in enumerate(header):
        n = _norm(v)
        if n and n not in norm_to_idx:
            norm_to_idx[n] = idx

    colmap: dict[str, int] = {}
    for field, names in aliases.items():
        for name in names:
            n = _norm(name)
            if n in norm_to_idx:
                colmap[field] = norm_to_idx[n]
                break

    missing = [f for f in required if f not in colmap]
    if missing:
        tried = "; ".join(f"{f}={aliases.get(f)}" for f in missing)
        raise ValueError(
            f"Sheet '{sheet_name}': required column(s) {missing} not found.\n"
            f"  Aliases tried: {tried}\n"
            f"  Headers seen : {sorted(norm_to_idx)}")
    return header_idx, colmap


def _cell(row, idx):
    if idx is None or idx >= len(row):
        return None
    return row[idx]


# ─── Read abstracts ───────────────────────────────────────────────────────────

def read_abstracts(conf: Conf) -> list[Talk]:
    wb = openpyxl.load_workbook(conf.INPUT_ABSTRACTS, read_only=True, data_only=True)

    ws = wb[conf.SCORES_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    header_idx, col = find_columns(
        rows, conf.COLUMN_ALIASES,
        required=["id", "title", "avg_score", "suggested_format", "submitter"],
        sheet_name=conf.SCORES_SHEET)

    id_re = re.compile(conf.ID_PATTERN)
    talks: list[Talk] = []
    for row in rows[header_idx + 1:]:
        raw_id = _cell(row, col["id"])
        if raw_id is None:
            continue
        abstract_id = str(raw_id)
        if not id_re.match(abstract_id):
            continue

        title     = str(_cell(row, col.get("title")) or "")
        sc        = _cell(row, col.get("avg_score"))
        avg_score = float(sc) if isinstance(sc, (int, float)) else 0.0
        suggested = str(_cell(row, col.get("suggested_format")) or "Poster")
        submitter = str(_cell(row, col.get("submitter")) or "")
        tag_val   = _cell(row, col.get("topic_tag"))
        topic     = str(tag_val).strip() if tag_val else ""

        spec = conf.FORMAT_MAP.get(suggested.strip().lower())
        if spec:
            talk_type, duration = spec["type"], int(spec["minutes"])
        else:
            talk_type = conf.FALLBACK_TYPE
            duration  = conf.duration_for_type(conf.FALLBACK_TYPE)
        if abstract_id in conf.INVITED_IDS:
            talk_type, duration = "Invited", conf.LONG_MIN

        talks.append(Talk(abstract_id, title, submitter, avg_score,
                          suggested, talk_type, duration, topic))

    # Second pass: poster-only submissions never scored (absent from scores sheet).
    seen_ids = {t.abstract_id for t in talks}
    ws_abs = wb[conf.ABSTRACTS_SHEET]
    arows = list(ws_abs.iter_rows(values_only=True))
    a_header_idx, acol = find_columns(
        arows, conf.ABSTRACTS_ALIASES,
        required=["id", "title"], sheet_name=conf.ABSTRACTS_SHEET)
    contrib_re = re.compile(conf.CONTRIBUTED_ID_PATTERN)
    for row in arows[a_header_idx + 1:]:
        raw_id = _cell(row, acol["id"])
        if raw_id is None:
            continue
        abstract_id = str(raw_id)
        if not contrib_re.match(abstract_id) or abstract_id in seen_ids:
            continue
        title     = str(_cell(row, acol.get("title")) or "")
        submitter = str(_cell(row, acol.get("submitter")) or "")
        talks.append(Talk(abstract_id, title, submitter, 0.0,
                          "Poster", "Poster", conf.SPARKLER_MIN, ""))

    return talks


# ─── Compute row indices from template structure ──────────────────────────────

def compute_slot_grid(conf: Conf) -> list[list[SessionSlot]]:
    """Sessions ordered session-major then day, matching Programme sheet rows."""
    session_info: list[tuple[int, int]] = []   # (first_talk_row, num_slots)
    r = 3
    for si, (_, sh, sm, eh, em) in enumerate(conf.SESSION_BLOCKS):
        r += 1  # session header
        r += 1  # chair row
        num_slots = ((eh * 60 + em) - (sh * 60 + sm)) // conf.SLOT_MIN
        session_info.append((r, num_slots))
        r += num_slots
        if conf.SESSION_HAS_BREAK_AFTER.get(si, False):
            r += 1

    sessions: list[list[SessionSlot]] = []
    for si, (first_row, num_slots) in enumerate(session_info):
        for di in range(conf.NUM_DAYS):
            slots = [
                SessionSlot(row=first_row + slot_i, col=di + 2,
                            day_index=di, session_index=si, slot_index=slot_i)
                for slot_i in range(num_slots)
            ]
            sessions.append(slots)
    return sessions


# ─── Scheduling algorithm ─────────────────────────────────────────────────────

def schedule(conf: Conf, talks: list[Talk],
             existing: dict[tuple, str]) -> tuple[dict, dict, list]:
    sessions_flat = compute_slot_grid(conf)
    n_long  = conf.LONG_MIN  // conf.SLOT_MIN
    n_short = conf.SHORT_MIN // conf.SLOT_MIN

    def is_available(slot: SessionSlot) -> bool:
        return not existing.get((slot.row, slot.col), "").strip()

    def _estimate_duration(cell_val: str) -> int:
        if cell_val == "__MERGED__":
            return 0
        if "[LONG" in cell_val or "[INV" in cell_val:
            return conf.LONG_MIN
        if "[SHORT" in cell_val:
            return conf.SHORT_MIN
        return conf.LONG_MIN

    used_min: dict[tuple, int] = {}
    assignments: dict[tuple, Talk] = {}
    spans:       dict[tuple, int]  = {}

    def next_free_slots(si: int, di: int, n: int) -> Optional[list[SessionSlot]]:
        group = next(
            (g for g in sessions_flat
             if g[0].session_index == si and g[0].day_index == di), None
        )
        if group is None:
            return None
        run: list[SessionSlot] = []
        for slot in group:
            run = (run + [slot]) if is_available(slot) else []
            if len(run) == n:
                return run
        return None

    def assign(slots: list[SessionSlot], talk: Talk) -> None:
        start = slots[0]
        assignments[(start.row, start.col)] = talk
        spans[(start.row, start.col)]       = len(slots)
        for slot in slots:
            existing[(slot.row, slot.col)] = talk.abstract_id
        used_min[(start.session_index, start.day_index)] = (
            used_min.get((start.session_index, start.day_index), 0) + talk.duration
        )

    rng = random.Random(conf.RANDOM_SEED)

    invited    = [t for t in talks if t.talk_type == "Invited"]
    long_oral  = [t for t in talks if t.talk_type == "Long"]
    short_oral = [t for t in talks if t.talk_type == "Short"]
    posters    = [t for t in talks if t.talk_type == "Poster"]

    oral_sessions: list[tuple[int, int]] = []
    for si in range(len(conf.SESSION_BLOCKS)):
        for di in range(conf.NUM_DAYS):
            group = next(g for g in sessions_flat
                         if g[0].session_index == si and g[0].day_index == di)
            oral_sessions.append((si, di))
            for slot in group:
                cv = existing.get((slot.row, slot.col), "").strip()
                if cv:
                    used_min[(si, di)] = (
                        used_min.get((si, di), 0) + _estimate_duration(cv)
                    )

    rng.shuffle(oral_sessions)

    # ── 0. Station-topic grouping: cluster tagged high-score talks together ──
    station_ids: set[str] = set()
    if conf.STATION_TAG:
        station_ids = {
            t.abstract_id for t in talks
            if t.topic.lower() == conf.STATION_TAG.lower()
            and t.avg_score >= conf.STATION_THRESHOLD
            and t.talk_type in ("Invited", "Long", "Short")
        }
    if station_ids:
        invited    = [t for t in invited    if t.abstract_id not in station_ids]
        long_oral  = [t for t in long_oral  if t.abstract_id not in station_ids]
        short_oral = [t for t in short_oral if t.abstract_id not in station_ids]

        st_talks   = [t for t in talks if t.abstract_id in station_ids]
        st_anchors = [t for t in st_talks if t.talk_type in ("Invited", "Long")]
        st_shorts  = [t for t in st_talks if t.talk_type == "Short"]
        st_queue   = st_anchors + st_shorts   # anchors land at the session front

        # Warn (don't fail) when a listed preferred session can't be used.
        for di, si in conf.STATION_SESSIONS:
            label = conf.SESSION_BLOCKS[si][0] if 0 <= si < len(conf.SESSION_BLOCKS) else "?"
            where = f"(day {di}, session {si} = {label})"
            if not (0 <= di < conf.NUM_DAYS) or not (0 <= si < len(conf.SESSION_BLOCKS)):
                print(f"  WARNING: station_sessions {where} is out of range "
                      f"(day 0-{conf.NUM_DAYS-1}, session 0-{len(conf.SESSION_BLOCKS)-1}) "
                      f"- skipped, talks will spill into other empty sessions")
            elif used_min.get((si, di), 0) != 0:
                print(f"  WARNING: station_sessions {where} is already occupied "
                      f"(fixed entry / sparkler) - skipped, talks will spill "
                      f"into other empty sessions")

        preferred      = [(si, di) for (di, si) in conf.STATION_SESSIONS]
        preferred_set  = set(preferred)
        station_order  = preferred + [s for s in oral_sessions
                                      if s not in preferred_set]

        for si, di in station_order:
            if not st_queue:
                break
            if used_min.get((si, di), 0) != 0:
                continue
            progressed = True
            while st_queue and progressed:
                progressed = False
                for idx, talk in enumerate(st_queue):
                    n = n_long if talk.talk_type in ("Invited", "Long") else n_short
                    slots = next_free_slots(si, di, n)
                    if slots:
                        assign(slots, st_queue.pop(idx))
                        progressed = True
                        break

    # ── 1. Invited talks: prefer fresh (empty) sessions ─────────────────────
    inv_queue = list(invited)
    for si, di in oral_sessions:
        if not inv_queue:
            break
        if used_min.get((si, di), 0) == 0:
            slots = next_free_slots(si, di, n_long)
            if slots:
                assign(slots, inv_queue.pop(0))

    # ── 2. Long orals: prefer fresh sessions then any with room ──────────────
    long_queue = list(long_oral)
    for si, di in oral_sessions:
        if not long_queue:
            break
        if used_min.get((si, di), 0) == 0:
            slots = next_free_slots(si, di, n_long)
            if slots:
                assign(slots, long_queue.pop(0))
    for si, di in oral_sessions:
        if not long_queue:
            break
        slots = next_free_slots(si, di, n_long)
        if slots:
            assign(slots, long_queue.pop(0))
    for t in long_queue:
        t.talk_type = "Poster (overflow)"
        t.duration  = conf.SPARKLER_MIN
        posters.append(t)

    # ── 3. Short orals: round-robin across sessions ─────────────────────────
    short_queue = list(short_oral)
    changed = True
    while changed and short_queue:
        changed = False
        for si, di in oral_sessions:
            if not short_queue:
                break
            slots = next_free_slots(si, di, n_short)
            if slots is not None:
                assign(slots, short_queue.pop(0))
                changed = True
    for t in short_queue:
        t.talk_type = "Poster (overflow)"
        t.duration  = conf.SPARKLER_MIN
        posters.append(t)

    # ── 4. Mark poster types ─────────────────────────────────────────────────
    for t in posters:
        t.talk_type = t.talk_type if "overflow" in t.talk_type else "Poster (sparkler)"

    # ── 5. Shuffle short talks within each session; invited/long stay at front ─
    for session_group in sessions_flat:
        filled_keys = [(s.row, s.col) for s in session_group
                       if (s.row, s.col) in assignments]
        if len(filled_keys) < 2:
            continue
        session_talks = [assignments[rc] for rc in filled_keys]
        anchors = [t for t in session_talks if t.talk_type in ("Invited", "Long")]
        shorts  = [t for t in session_talks if t.talk_type not in ("Invited", "Long")]
        rng.shuffle(shorts)
        for rc, talk in zip(filled_keys, anchors + shorts):
            assignments[rc] = talk

    return assignments, spans, posters


# ─── Write to Excel ───────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

COL_INVITED  = "BDD7EE"
COL_LONG     = "9DC3E6"
COL_SHORT    = "E2EFDA"
COL_POSTER   = "FCE4D6"
COL_OVERFLOW = "F4CCCC"

TYPE_COLOR = {
    "Invited":           COL_INVITED,
    "Long":              COL_LONG,
    "Short":             COL_SHORT,
    "Poster":            COL_POSTER,
    "Poster (sparkler)": COL_POSTER,
    "Poster (overflow)": COL_OVERFLOW,
}

TYPE_LABEL = {
    "Invited": "INV 30'",
    "Long":    "LONG 30'",
    "Short":   "SHORT 15'",
    "Poster":  "POSTER",
    "Poster (sparkler)": "POSTER",
    "Poster (overflow)": "POSTER",
}


def write_assignments(conf: Conf, assignments: dict[tuple, Talk],
                      spans: dict[tuple, int], posters: list[Talk]):
    wb = openpyxl.load_workbook(conf.SCHEDULE_FILE)
    ws = wb["Programme"]

    existing_merges = [
        (mr.min_row, mr.min_col, mr.max_row, mr.max_col)
        for mr in ws.merged_cells.ranges
    ]

    for (row, col), talk in sorted(assignments.items()):
        cell = ws.cell(row, col)
        if cell.value and str(cell.value).strip():
            continue

        n_rows = spans.get((row, col), 1)
        if n_rows > 1:
            already = any(
                mr[0] <= row <= mr[2] and mr[1] <= col <= mr[3]
                for mr in existing_merges
            )
            if not already:
                ws.merge_cells(start_row=row, start_column=col,
                               end_row=row + n_rows - 1, end_column=col)

        cell   = ws.cell(row, col)
        label  = TYPE_LABEL.get(talk.talk_type, "")
        title  = textwrap.shorten(talk.title, width=70, placeholder="…")
        name   = talk.submitter.split("\n")[0].split(",")[0].strip()
        name   = textwrap.shorten(name, width=35, placeholder="…")

        cell.value     = f"[{label}] {talk.abstract_id}\n{title}\n{name}"
        cell.fill      = _fill(TYPE_COLOR.get(talk.talk_type, "FFFFFF"))
        cell.font      = Font(size=8, bold=(talk.talk_type in ("Invited", "Long")))
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        cell.border    = _thin_border()

    # ── Write Posters sheet ───────────────────────────────────────────────────
    wp = wb["Posters"]
    wp.delete_rows(2, wp.max_row)

    sparkler_note = ""
    if conf.SPARKLER_SESSIONS:
        di, si = conf.SPARKLER_SESSIONS[0][0], conf.SPARKLER_SESSIONS[0][1]
        d = conf.CONF_START + timedelta(days=di)
        sl = conf.SESSION_BLOCKS[si]
        sparkler_note = (f"{d.strftime('%A %d %b')}  |  "
                         f"{sl[0]}  {sl[1]:02d}:{sl[2]:02d}–{sl[3]:02d}:{sl[4]:02d}")

    wp.cell(1, 6, f"Sparkler session: {sparkler_note}").font = Font(
        italic=True, color="7F6000")

    for i, talk in enumerate(posters, start=2):
        wp.cell(i, 1, talk.abstract_id)
        wp.cell(i, 2, talk.title)
        wp.cell(i, 3, round(talk.avg_score, 2))
        wp.cell(i, 4, talk.talk_type)
        wp.cell(i, 5, talk.submitter.split("\n")[0].split(",")[0].strip())
        for c in range(1, 6):
            wp.cell(i, c).fill   = _fill(TYPE_COLOR.get(talk.talk_type, "FFFFFF"))
            wp.cell(i, c).border = _thin_border()
            wp.cell(i, c).font   = Font(size=9)

    wb.save(conf.SCHEDULE_FILE)


# ─── Main ─────────────────────────────────────────────────────────────────────

def main(conf: Conf, dry_run: bool = False):
    seed_str = str(conf.RANDOM_SEED) if conf.RANDOM_SEED is not None else "random (no seed)"
    print(f"Random seed: {seed_str}  (change random_seed in the YAML for a different layout)")
    print(f"Reading abstracts from: {conf.INPUT_ABSTRACTS}")
    talks = read_abstracts(conf)
    print(f"  {len(talks)} abstracts loaded")

    types = {}
    for t in talks:
        types[t.talk_type] = types.get(t.talk_type, 0) + 1
    for k, v in sorted(types.items()):
        print(f"    {k:<20} {v}")

    print(f"\nReading schedule template: {conf.SCHEDULE_FILE}")
    wb = openpyxl.load_workbook(conf.SCHEDULE_FILE, data_only=True)
    ws = wb["Programme"]

    existing: dict[tuple, str] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and str(cell.value).strip():
                existing[(cell.row, cell.column)] = str(cell.value)

    for mr in ws.merged_cells.ranges:
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if (r, c) not in existing:
                    existing[(r, c)] = "__MERGED__"

    print(f"  {len(existing)} pre-filled cells (will not be overwritten)")

    placed_ids: set[str] = set()
    id_re = re.compile(conf.ID_PATTERN)
    for cell_val in existing.values():
        for part in cell_val.split():
            if id_re.match(part):
                placed_ids.add(part)
                break
    if placed_ids:
        before = len(talks)
        talks = [t for t in talks if t.abstract_id not in placed_ids]
        print(f"  {len(placed_ids)} talks already in schedule — {before - len(talks)} filtered out")

    if conf.STATION_TAG:
        station_q = [t for t in talks
                     if t.topic.lower() == conf.STATION_TAG.lower()
                     and t.avg_score >= conf.STATION_THRESHOLD
                     and t.talk_type in ("Invited", "Long", "Short")]
        if station_q:
            print(f"\n'{conf.STATION_TAG}' talks grouped (score >= {conf.STATION_THRESHOLD}): "
                  f"{len(station_q)}")
            for t in station_q:
                print(f"    {t.abstract_id}  {t.avg_score:.2f}  {t.talk_type:<7} "
                      f"{t.title[:50]}")

    assignments, spans, posters = schedule(conf, talks, existing)
    print(f"\nScheduling result:")
    print(f"  Oral talk slots filled : {len(assignments)}")
    print(f"  Posters/sparkler       : {len(posters)}")

    if posters:
        print(f"\n  Poster talks (score order):")
        for t in posters[:20]:
            print(f"    {t.abstract_id}  {t.avg_score:.2f}  {t.title[:55]}")
        if len(posters) > 20:
            print(f"    ... and {len(posters)-20} more")

    if dry_run:
        print("\n[dry-run] No changes written.")
        return

    write_assignments(conf, assignments, spans, posters)
    print(f"\nSaved: {conf.SCHEDULE_FILE}")

    sessions_flat = compute_slot_grid(conf)
    print("\nSession utilisation:")
    print(f"  {'Day':<14}  {'Session':<14}  {'Talks':<6}  {'Notes'}")
    sparkler_set = {(e[0], e[1]) for e in conf.SPARKLER_SESSIONS}
    assigned_positions = set(assignments.keys())
    for si, (s_label, sh, sm, eh, em) in enumerate(conf.SESSION_BLOCKS):
        for di in range(conf.NUM_DAYS):
            d = conf.CONF_START + timedelta(days=di)
            group = [g for g in sessions_flat
                     if g[0].session_index == si and g[0].day_index == di][0]
            filled = sum(1 for s in group if (s.row, s.col) in assigned_positions)
            notes = ""
            if (di, si) in sparkler_set:
                notes = "Sparkler"
            print(f"  {d.strftime('%a %d %b'):<14}  {s_label:<14}  {filled:<6}  {notes}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Fill a schedule grid with scored abstracts from a YAML config.")
    parser.add_argument("--config", default="conference.yaml",
                        help="Path to the conference YAML (default: conference.yaml)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Print the plan without writing to the workbook")
    args = parser.parse_args()
    main(load(args.config), dry_run=args.dry_run)
