#!/usr/bin/env python3
"""
generate_schedule.py  —  Create a blank conference schedule template.

Every conference-specific setting
comes from a YAML file (default conference.yaml) instead of an in-file CONFIG
block, so the same script works for any conference.

    python generate_schedule.py                       # uses conference.yaml
    python generate_schedule.py --config other.yaml

The output workbook (schedule_file in the YAML) has:
  - Programme  — visual grid (days x session blocks)
  - Posters    — poster / sparkler talk list (populated by fill_schedule.py)
  - _META      — hidden sheet recording row/col structure for the fill step

fill_schedule.py fills every cell that is currently EMPTY.  Pre-fill any
cell you want to "lock" via fixed_entries and the fill script never overwrites it.
"""

import argparse
from datetime import timedelta

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from schedule_config import load, Conf


# ── helpers ──────────────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_bottom():
    thick = Side(style="medium", color="666666")
    thin  = Side(style="thin",  color="AAAAAA")
    return Border(left=thin, right=thin, top=thin, bottom=thick)

COL_CONF_HDR   = "1F4E79"   # dark navy  - conference title
COL_DAY_HDR    = "2E75B6"   # mid blue   - day header
COL_SESS_HDR   = "D6E4F0"   # light blue - session header
COL_BREAK_HDR  = "F2F2F2"   # light grey - break rows
COL_SPARKLER   = "FFE699"   # amber      - sparkler session cell
COL_CHAIR      = "EAF0FB"   # pale blue-grey - chair row
COL_FIXED      = "C6EFCE"   # light green - pre-fixed entries
COL_EMPTY      = "FFFFFF"   # white      - empty talk slots

FONT_CONF  = Font(bold=True, color="FFFFFF", size=14)
FONT_DAY   = Font(bold=True, color="FFFFFF", size=11)
FONT_SESS  = Font(bold=True, color="1F4E79", size=10)
FONT_BREAK = Font(italic=True, color="666666", size=9)
FONT_FIXED = Font(bold=True, color="375623", size=9)
FONT_NORM  = Font(size=9)
FONT_TIME  = Font(bold=True, size=8, color="1F4E79")

ALIGN_CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)


def build_schedule(conf: Conf):
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Programme"

    # Build lookup: (di, si) -> list of (text, start_slot, end_slot_exclusive)
    # Multiple entries for the same (day, session) are placed sequentially unless
    # an explicit start_slot is given.
    fixed_map: dict[tuple, list[tuple]] = {}
    for e in conf.FIXED_ENTRIES:
        di, si, txt = e[0], e[1], e[2]
        n = e[3] if len(e) > 3 else 1
        key = (di, si)
        if len(e) > 4:
            start = e[4]                                          # explicit start slot
        else:
            start = fixed_map[key][-1][2] if key in fixed_map else 0   # auto-stack
        fixed_map.setdefault(key, []).append((txt, start, start + n))
    sparkler_set = {(e[0], e[1]) for e in conf.SPARKLER_SESSIONS}
    # (start_slot, n_slots); None means "resolve from session length inside loop"
    sparkler_map = {(e[0], e[1]): (e[2] if len(e) > 2 else 0,
                                   e[3] if len(e) > 3 else None)
                    for e in conf.SPARKLER_SESSIONS}
    break_after = {b[1]: b[0] for b in conf.BREAKS}   # session_index -> break label

    # ── Validate fixed / sparkler placement before drawing anything ───────────
    session_slots = [((eh * 60 + em) - (sh * 60 + sm)) // conf.SLOT_MIN
                     for _, sh, sm, eh, em in conf.SESSION_BLOCKS]
    problems: list[str] = []
    for (di, si), entries in fixed_map.items():
        if not (0 <= di < conf.NUM_DAYS):
            problems.append(f"fixed entry day_index {di} out of range "
                            f"(0-{conf.NUM_DAYS-1})")
            continue
        if not (0 <= si < len(conf.SESSION_BLOCKS)):
            problems.append(f"fixed entry session_index {si} out of range "
                            f"(0-{len(conf.SESSION_BLOCKS)-1})")
            continue
        n_slots = session_slots[si]
        spans = [(txt, s, e) for txt, s, e in entries]
        if (di, si) in sparkler_map:
            sp_start, sp_raw = sparkler_map[(di, si)]
            sp_n = sp_raw if sp_raw is not None else (n_slots - sp_start)
            spans.append(("Poster Sparkler Session", sp_start, sp_start + sp_n))
        for txt, s, e in spans:
            if s < 0 or e > n_slots:
                problems.append(
                    f"'{txt}' (day {di}, session {si}) spans slots {s}-{e-1} "
                    f"but session has only {n_slots} slots (0-{n_slots-1})")
        ordered = sorted(spans, key=lambda x: x[1])
        for (t1, s1, e1), (t2, s2, e2) in zip(ordered, ordered[1:]):
            if s2 < e1:
                problems.append(
                    f"overlap in day {di}, session {si}: '{t1}' (slots {s1}-{e1-1}) "
                    f"and '{t2}' (slots {s2}-{e2-1})")
    if problems:
        raise ValueError("Schedule placement errors:\n  - " + "\n  - ".join(problems))

    days = [conf.CONF_START + timedelta(days=i) for i in range(conf.NUM_DAYS)]

    # Column layout: col 1 = session label, cols 2..(NUM_DAYS+1) = days
    NUM_COLS = conf.NUM_DAYS + 1

    # ── Row 1 : conference title ──────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
    hdr = ws.cell(1, 1, f"{conf.CONF_NAME}  ·  {conf.CONF_VENUE}  ·  "
                         f"{days[0].strftime('%d %b')} – {days[-1].strftime('%d %b %Y')}")
    hdr.font      = FONT_CONF
    hdr.fill      = _fill(COL_CONF_HDR)
    hdr.alignment = ALIGN_CTR
    ws.row_dimensions[1].height = 28

    # ── Row 2 : day headers ───────────────────────────────────────────────────
    ws.cell(2, 1, "Session").font = FONT_DAY
    ws.cell(2, 1).fill = _fill(COL_DAY_HDR)
    ws.cell(2, 1).alignment = ALIGN_CTR
    for ci, d in enumerate(days, start=2):
        c = ws.cell(2, ci, d.strftime("%A\n%d %b %Y"))
        c.font      = FONT_DAY
        c.fill      = _fill(COL_DAY_HDR)
        c.alignment = ALIGN_CTR
    ws.row_dimensions[2].height = 32

    current_row = 3

    for si, (s_label, sh, sm, eh, em) in enumerate(conf.SESSION_BLOCKS):
        duration_min = (eh * 60 + em) - (sh * 60 + sm)
        num_slots    = duration_min // conf.SLOT_MIN
        time_range   = f"{sh:02d}:{sm:02d} – {eh:02d}:{em:02d}"

        # ── Session header row ────────────────────────────────────────────────
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row,   end_column=NUM_COLS)
        sh_cell = ws.cell(current_row, 1, f"{s_label}   {time_range}")
        sh_cell.font      = FONT_SESS
        sh_cell.fill      = _fill(COL_SESS_HDR)
        sh_cell.alignment = ALIGN_CTR
        sh_cell.border    = _thick_bottom()
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        # ── Chair row ─────────────────────────────────────────────────────────
        lbl = ws.cell(current_row, 1, "Chair")
        lbl.font      = Font(bold=True, size=8, color="1F4E79")
        lbl.alignment = Alignment(horizontal="center", vertical="center")
        lbl.fill      = _fill(COL_CHAIR)
        lbl.border    = _thin_border()
        for di in range(conf.NUM_DAYS):
            c = ws.cell(current_row, di + 2)
            c.fill   = _fill(COL_CHAIR)
            c.border = _thin_border()
            c.font   = Font(italic=True, size=8, color="888888")
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        # ── Time-slot rows: one row per SLOT_MIN minutes ──────────────────────
        first_slot_row = current_row   # remember for post-loop merging
        for slot in range(num_slots):
            slot_h, slot_m = divmod(sh * 60 + sm + slot * conf.SLOT_MIN, 60)
            tc = ws.cell(current_row, 1, f"{slot_h:02d}:{slot_m:02d}")
            tc.font      = FONT_TIME
            tc.alignment = Alignment(horizontal="center", vertical="center")
            tc.border    = _thin_border()

            for di, d in enumerate(days):
                col = di + 2
                key = (di, si)
                c   = ws.cell(current_row, col)
                c.border    = _thin_border()
                c.alignment = ALIGN_L
                c.font      = FONT_NORM

                if key in sparkler_set:
                    sp_start, sp_raw = sparkler_map[key]
                    sp_n = sp_raw if sp_raw is not None else (num_slots - sp_start)
                    if sp_start <= slot < sp_start + sp_n:
                        c.fill = _fill(COL_SPARKLER)
                    else:
                        c.fill = _fill(COL_EMPTY)
                elif key in fixed_map:
                    entry = next(
                        (e for e in fixed_map[key] if e[1] <= slot < e[2]),
                        None,
                    )
                    if entry is not None:
                        etxt, estart, eend = entry
                        if slot == estart and (eend - estart) == 1:
                            c = ws.cell(current_row, col, etxt)
                            c.font = FONT_FIXED
                        c.fill = _fill(COL_FIXED)
                    else:
                        c.fill = _fill(COL_EMPTY)
                else:
                    c.fill = _fill(COL_EMPTY)

            ws.row_dimensions[current_row].height = 42
            current_row += 1

        # ── Merge multi-slot fixed entries after all rows are written ─────────
        for di in range(conf.NUM_DAYS):
            key = (di, si)
            if key not in fixed_map:
                continue
            for etxt, estart, eend in fixed_map[key]:
                n_fix = eend - estart
                if n_fix <= 1:
                    continue
                col = di + 2
                start_row = first_slot_row + estart
                ws.merge_cells(start_row=start_row, start_column=col,
                               end_row=start_row + n_fix - 1, end_column=col)
                mc = ws.cell(start_row, col, etxt)
                mc.font      = FONT_FIXED
                mc.fill      = _fill(COL_FIXED)
                mc.alignment = ALIGN_CTR
                mc.border    = _thin_border()

        # ── Merge sparkler cells (full or partial session) ────────────────────
        for di in range(conf.NUM_DAYS):
            if (di, si) not in sparkler_set:
                continue
            sp_start, sp_raw = sparkler_map[(di, si)]
            sp_n = sp_raw if sp_raw is not None else (num_slots - sp_start)
            col = di + 2
            ws.merge_cells(start_row=first_slot_row + sp_start, start_column=col,
                           end_row=first_slot_row + sp_start + sp_n - 1, end_column=col)
            mc = ws.cell(first_slot_row + sp_start, col, "Poster Sparkler Session")
            mc.font      = Font(bold=True, color="7F6000", size=9)
            mc.fill      = _fill(COL_SPARKLER)
            mc.alignment = ALIGN_CTR
            mc.border    = _thin_border()

        # ── Break row (if applicable) ─────────────────────────────────────────
        if si in break_after:
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row,   end_column=NUM_COLS)
            br = ws.cell(current_row, 1, break_after[si])
            br.font      = FONT_BREAK
            br.fill      = _fill(COL_BREAK_HDR)
            br.alignment = ALIGN_CTR
            ws.row_dimensions[current_row].height = 16
            current_row += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 8
    for ci in range(2, NUM_COLS + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 32

    # ── Freeze panes: keep header + day row visible ───────────────────────────
    ws.freeze_panes = "B3"

    # ── Posters sheet (filled later) ──────────────────────────────────────────
    wp = wb.create_sheet("Posters")
    headers = ["Abstract ID", "Title", "Avg Score", "Type", "Submitter"]
    for ci, h in enumerate(headers, 1):
        c = wp.cell(1, ci, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = _fill(COL_DAY_HDR)
        c.alignment = ALIGN_CTR
    for ci, w in enumerate([12, 70, 10, 18, 40], 1):
        wp.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    # ── META sheet: records the schedule structure for fill_schedule.py ─
    wm = wb.create_sheet("_META")
    wm.sheet_state = "hidden"

    meta_row = 1
    current_row = 3
    for si, (s_label, sh, sm, eh, em) in enumerate(conf.SESSION_BLOCKS):
        duration_min = (eh * 60 + em) - (sh * 60 + sm)
        num_slots    = duration_min // conf.SLOT_MIN
        first_talk = current_row + 2       # skip session header + chair row
        last_talk  = first_talk + num_slots - 1
        wm.cell(meta_row, 1, si)
        wm.cell(meta_row, 2, s_label)
        wm.cell(meta_row, 3, first_talk)
        wm.cell(meta_row, 4, last_talk)
        meta_row += 1
        current_row += 2 + num_slots + (1 if si in break_after else 0)

    for di in range(conf.NUM_DAYS):
        wm.cell(meta_row, 1, di)
        wm.cell(meta_row, 2, (conf.CONF_START + timedelta(days=di)).isoformat())
        wm.cell(meta_row, 3, di + 2)       # Excel column number
        meta_row += 1

    wb.save(conf.OUTPUT_FILE)
    print(f"Created: {conf.OUTPUT_FILE}")
    total_slots = sum(((eh*60+em)-(sh*60+sm))//conf.SLOT_MIN
                      for _, sh, sm, eh, em in conf.SESSION_BLOCKS) * conf.NUM_DAYS
    print(f"  {conf.NUM_DAYS} days  x  {len(conf.SESSION_BLOCKS)} sessions  x  "
          f"{conf.SLOT_MIN}-min rows  =  {total_slots} total time slots")
    print(f"  Fixed entries: {len(conf.FIXED_ENTRIES)}")
    print(f"  Sparkler sessions: {len(conf.SPARKLER_SESSIONS)}")
    print(f"\nNext: run  python fill_schedule.py --config <your.yaml>")


def main():
    parser = argparse.ArgumentParser(
        description="Create a blank conference schedule template from a YAML config.")
    parser.add_argument("--config", default="conference.yaml",
                        help="Path to the conference YAML (default: conference.yaml)")
    args = parser.parse_args()
    conf = load(args.config)
    build_schedule(conf)


if __name__ == "__main__":
    main()
